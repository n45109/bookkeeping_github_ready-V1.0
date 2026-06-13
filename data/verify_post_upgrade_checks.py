import json
from pathlib import Path
import shutil
import sqlite3
import sys
from datetime import datetime

BASE_DIR = Path(__file__).resolve().parents[1]
if str(BASE_DIR) not in sys.path:
    sys.path.insert(0, str(BASE_DIR))

import database

from database import (
    REQUIRED_RECORD_COLUMNS,
    REQUIRED_TABLES,
    REQUIRED_USER_COLUMNS,
    assign_user_organization,
    column_exists,
    create_organization,
    create_user,
    export_to_excel,
    get_all_records,
    get_balance,
    get_balance_history,
    get_conn,
    get_current_schema_version,
    get_setting,
    get_user_by_username,
    has_setting,
    init_db,
    load_records,
    save_records,
    schema_needs_migration,
    set_setting,
    table_exists,
)


def prepare_temp_database():
    temp_root = BASE_DIR / "data" / "_verify_tmp"
    temp_root.mkdir(parents=True, exist_ok=True)
    run_id = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
    temp_dir = temp_root / f"post_upgrade_{run_id}"
    temp_dir.mkdir(parents=True, exist_ok=True)
    temp_db_path = temp_dir / "bookkeeping_verify.db"
    source_db_path = BASE_DIR / "data" / "bookkeeping.db"

    if source_db_path.exists():
        shutil.copy2(source_db_path, temp_db_path)
        for suffix in (".db-wal", ".db-shm"):
            source_sidecar = source_db_path.with_suffix(suffix)
            target_sidecar = temp_db_path.with_suffix(suffix)
            if source_sidecar.exists():
                shutil.copy2(source_sidecar, target_sidecar)
    else:
        conn = sqlite3.connect(temp_db_path)
        conn.close()

    database.DATA_DIR = temp_dir
    database.DB_PATH = temp_db_path
    database.BACKUP_DIR = temp_dir / "backups"
    database.MIGRATION_BACKUP_DONE = False
    return temp_db_path


def ensure_user(
    username: str,
    password: str,
    display_name: str,
    organization_id: int,
    role: str,
    is_admin: bool = False,
):
    user = get_user_by_username(username)
    if user is None:
        user = create_user(username, password, display_name, organization_id, role)
    return assign_user_organization(
        user["id"],
        organization_id,
        role=role,
        is_admin=is_admin,
    )


def make_record(
    *,
    date: str,
    project_name: str,
    applicant: str,
    receiver: str,
    income: float = 0,
    expense: float = 0,
    owner_user_id: int,
):
    return {
        "日期": date,
        "项目名称": project_name,
        "申请人": applicant,
        "转款人": "",
        "收款人": receiver,
        "用途": "post-upgrade verification",
        "收入": income,
        "支出": expense,
        "可退回": "否",
        "预计退回": "",
        "状态": "",
        "备注": f"{project_name} verification record",
        "owner_user_id": owner_user_id,
    }


def verify_schema_state():
    conn = get_conn()
    try:
        tables_ok = all(table_exists(conn, table_name) for table_name in REQUIRED_TABLES)
        user_columns_ok = all(column_exists(conn, "users", name) for name in REQUIRED_USER_COLUMNS)
        record_columns_ok = all(column_exists(conn, "records", name) for name in REQUIRED_RECORD_COLUMNS)
        schema_version = get_current_schema_version(conn)
        needs_migration = schema_needs_migration(conn)
    finally:
        conn.close()
    return {
        "tables_ok": tables_ok,
        "user_columns_ok": user_columns_ok,
        "record_columns_ok": record_columns_ok,
        "schema_version": schema_version,
        "needs_migration": needs_migration,
    }


def verify_org_isolation():
    org_a = create_organization("Post Upgrade Org A")
    org_b = create_organization("Post Upgrade Org B")

    boss_a = ensure_user("verify_boss_a", "123456", "Verify Boss A", org_a["id"], "boss")
    staff_a = ensure_user("verify_staff_a", "123456", "Verify Staff A", org_a["id"], "staff")
    boss_b = ensure_user("verify_boss_b", "123456", "Verify Boss B", org_b["id"], "boss")
    staff_b = ensure_user("verify_staff_b", "123456", "Verify Staff B", org_b["id"], "staff")

    save_records([make_record(date="2026-06-07", project_name="PUG-A-Boss", applicant="Boss A", receiver="A", expense=101, owner_user_id=boss_a["id"])], boss_a["id"])
    save_records([make_record(date="2026-06-07", project_name="PUG-A-Staff", applicant="Staff A", receiver="A", expense=102, owner_user_id=staff_a["id"])], staff_a["id"])
    save_records([make_record(date="2026-06-07", project_name="PUG-B-Boss", applicant="Boss B", receiver="B", expense=201, owner_user_id=boss_b["id"])], boss_b["id"])
    save_records([make_record(date="2026-06-07", project_name="PUG-B-Staff", applicant="Staff B", receiver="B", expense=202, owner_user_id=staff_b["id"])], staff_b["id"])

    def visible_projects(user: dict):
        return [
            row["project_name"]
            for row in get_all_records(user)["records"]
            if row["project_name"].startswith("PUG-")
        ]

    result = {
        "boss_a_projects": visible_projects(boss_a),
        "staff_a_projects": visible_projects(staff_a),
        "boss_b_projects": visible_projects(boss_b),
        "staff_b_projects": visible_projects(staff_b),
    }
    checks = {
        "boss_a_sees_org_a_only": sorted(result["boss_a_projects"]) == ["PUG-A-Boss", "PUG-A-Staff"],
        "staff_a_sees_self_only": result["staff_a_projects"] == ["PUG-A-Staff"],
        "boss_b_sees_org_b_only": sorted(result["boss_b_projects"]) == ["PUG-B-Boss", "PUG-B-Staff"],
        "staff_b_sees_self_only": result["staff_b_projects"] == ["PUG-B-Staff"],
    }
    return {
        "result": result,
        "checks": checks,
        "all_passed": all(checks.values()),
    }


def verify_initial_balance_zero():
    org = create_organization("Post Upgrade Balance Org")
    user = ensure_user("verify_balance_zero", "123456", "Verify Balance Zero", org["id"], "staff")
    set_setting(user["id"], "initial_balance", "0")

    loaded = load_records(user)
    checks = {
        "initial_balance_marked_as_set": loaded["initial_balance_set"] is True,
        "has_setting_reports_true": has_setting(user["id"], "initial_balance") is True,
        "stored_value_is_zero": get_setting(user["id"], "initial_balance") == "0",
        "balance_starts_at_zero": float(loaded["balance"]) == 0.0,
    }
    return {
        "user_id": user["id"],
        "checks": checks,
        "all_passed": all(checks.values()),
    }


def verify_read_write_smoke():
    org = create_organization("Post Upgrade Smoke Org")
    staff = ensure_user("verify_smoke_staff", "123456", "Verify Smoke Staff", org["id"], "staff")

    before = load_records(staff)
    save_records(
        [
            make_record(
                date="2026-06-09",
                project_name="Smoke-WriteRead",
                applicant="Smoke Staff",
                receiver="Receiver",
                expense=88,
                owner_user_id=staff["id"],
            )
        ],
        staff["id"],
    )
    after = load_records(staff)
    smoke_records = [row for row in after["records"] if row["project_name"] == "Smoke-WriteRead"]
    checks = {
        "before_load_is_readable": isinstance(before["records"], list),
        "after_load_is_readable": isinstance(after["records"], list),
        "written_record_can_be_read_back": len(smoke_records) == 1,
        "staff_scope_still_limits_to_own_record": smoke_records[0]["owner_user_id"] == staff["id"] if smoke_records else False,
    }
    return {
        "user_id": staff["id"],
        "checks": checks,
        "all_passed": all(checks.values()),
    }


def verify_balance_order_and_consistency():
    org = create_organization("Post Upgrade Timeline Org")
    boss = ensure_user("verify_timeline_boss", "123456", "Verify Timeline Boss", org["id"], "boss")

    set_setting(boss["id"], "initial_balance", "1000")
    save_records(
        [
            make_record(date="2026-06-10", project_name="Timeline-Late", applicant="Boss", receiver="A", expense=100, owner_user_id=boss["id"]),
            make_record(date="2026-06-08", project_name="Timeline-Early", applicant="Boss", receiver="A", expense=50, owner_user_id=boss["id"]),
            make_record(date="2026-06-10", project_name="Timeline-SameDay", applicant="Boss", receiver="A", income=20, owner_user_id=boss["id"]),
        ],
        boss["id"],
    )

    history = [
        row for row in get_balance_history(boss)
        if row["project_name"].startswith("Timeline-")
    ]
    ordered_projects = [row["project_name"] for row in history]
    ordered_balances = [row["balance"] for row in history]
    expected_projects = ["Timeline-Early", "Timeline-Late", "Timeline-SameDay"]
    expected_balances = [950.0, 850.0, 870.0]

    loaded = load_records(boss)
    exported_source = [
        row["project_name"]
        for row in loaded["records"]
        if row["project_name"].startswith("Timeline-")
    ]
    checks = {
        "projects_order_matches_date_then_id": ordered_projects == expected_projects,
        "balance_progression_matches_expected": ordered_balances == expected_balances,
        "load_records_uses_same_order": exported_source == expected_projects,
        "final_balance_matches_total": float(get_balance(boss)) == 870.0,
    }
    return {
        "ordered_projects": ordered_projects,
        "ordered_balances": ordered_balances,
        "checks": checks,
        "all_passed": all(checks.values()),
    }


def verify_export_backend():
    from openpyxl import load_workbook

    org = create_organization("Post Upgrade Export Org")
    boss = ensure_user("verify_export_boss", "123456", "Verify Export Boss", org["id"], "boss")
    set_setting(boss["id"], "initial_balance", "300")
    save_records(
        [
            make_record(date="2026-06-09", project_name="Export-Check", applicant="Boss", receiver="Receiver", expense=10, owner_user_id=boss["id"])
        ],
        boss["id"],
    )

    path = Path(export_to_excel(boss))
    workbook = load_workbook(path)
    sheet = workbook.active
    title_row = [sheet.cell(row=1, column=index).value for index in range(1, 14)]
    initial_balance_label = sheet.cell(row=2, column=1).value
    initial_balance_value = sheet.cell(row=2, column=9).value
    first_record_name = sheet.cell(row=3, column=2).value
    workbook.close()
    checks = {
        "export_file_exists": path.exists(),
        "export_is_xlsx": path.suffix.lower() == ".xlsx",
        "export_file_non_empty": path.exists() and path.stat().st_size > 0,
        "export_header_is_readable": title_row[:3] == ["日期", "项目名称", "申请人"],
        "initial_balance_row_is_present": initial_balance_label == "初始余额" and float(initial_balance_value) == 300.0,
        "first_record_is_present": first_record_name == "Export-Check",
    }
    return {
        "path": str(path),
        "checks": checks,
        "all_passed": all(checks.values()),
    }


def main():
    database_path = prepare_temp_database()
    init_db()

    schema_result = verify_schema_state()
    org_result = verify_org_isolation()
    zero_balance_result = verify_initial_balance_zero()
    smoke_result = verify_read_write_smoke()
    timeline_result = verify_balance_order_and_consistency()
    export_result = verify_export_backend()

    all_checks = {
        "schema_ready": (
            schema_result["tables_ok"]
            and schema_result["user_columns_ok"]
            and schema_result["record_columns_ok"]
            and not schema_result["needs_migration"]
        ),
        "organization_isolation": org_result["all_passed"],
        "initial_balance_zero_handled": zero_balance_result["all_passed"],
        "read_write_smoke": smoke_result["all_passed"],
        "running_balance_rule": timeline_result["all_passed"],
        "export_backend_ready": export_result["all_passed"],
    }

    print(
        json.dumps(
            {
                "database_used": str(database_path),
                "temp_directory": str(database.DATA_DIR),
                "schema": schema_result,
                "organization_isolation": org_result,
                "initial_balance_zero": zero_balance_result,
                "read_write_smoke": smoke_result,
                "running_balance_rule": timeline_result,
                "export_backend": export_result,
                "summary_checks": all_checks,
                "all_passed": all(all_checks.values()),
            },
            ensure_ascii=False,
            indent=2,
        )
    )


if __name__ == "__main__":
    main()
