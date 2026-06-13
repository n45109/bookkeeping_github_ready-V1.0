"""
数据库模块 - SQLite + 多用户隔离 + 可迁移结构
"""
import hashlib
import secrets
import sqlite3
import time
from datetime import datetime, timedelta
from pathlib import Path
import shutil

BASE_DIR = Path(__file__).parent
DATA_DIR = BASE_DIR / "data"
DB_PATH = DATA_DIR / "bookkeeping.db"
BACKUP_DIR = DATA_DIR / "backups"
SCHEMA_VERSION = 4
ROLE_ADMIN = "admin"
ROLE_BOSS = "boss"
ROLE_STAFF = "staff"
VALID_ROLES = {ROLE_ADMIN, ROLE_BOSS, ROLE_STAFF}
DEFAULT_ORGANIZATION_NAME = "神秘组织"
LOCK_RETRY_DELAYS = (0.15, 0.3, 0.6)
REQUIRED_TABLES = (
    "users",
    "sessions",
    "organizations",
    "records",
    "audit_events",
    "settings",
    "schema_migrations",
)
REQUIRED_USER_COLUMNS = ("organization_id", "role")
REQUIRED_RECORD_COLUMNS = ("organization_id", "created_by_user_id", "owner_user_id")
MIGRATION_BACKUP_DONE = False


def get_conn():
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(str(DB_PATH), timeout=10)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA foreign_keys=ON")
    conn.execute("PRAGMA busy_timeout = 10000")
    ensure_db_ready(conn)
    return conn


def run_write(operation, *, retries=LOCK_RETRY_DELAYS):
    last_error = None
    for attempt, delay in enumerate((0, *retries), start=1):
        conn = get_conn()
        try:
            result = operation(conn)
            conn.commit()
            conn.close()
            return result
        except sqlite3.OperationalError as exc:
            conn.close()
            last_error = exc
            if "locked" not in str(exc).lower() or attempt > len(retries):
                raise
            time.sleep(delay)
        except Exception:
            conn.close()
            raise
    if last_error:
        raise last_error


def table_exists(conn, table_name: str) -> bool:
    row = conn.execute(
        "SELECT name FROM sqlite_master WHERE type = 'table' AND name = ?",
        (table_name,),
    ).fetchone()
    return row is not None


def column_exists(conn, table_name: str, column_name: str) -> bool:
    rows = conn.execute(f"PRAGMA table_info({table_name})").fetchall()
    return any(row[1] == column_name for row in rows)


def ensure_column(conn, table_name: str, column_name: str, ddl: str):
    if not column_exists(conn, table_name, column_name):
        conn.execute(f"ALTER TABLE {table_name} ADD COLUMN {ddl}")


def has_required_tables(conn) -> bool:
    return all(table_exists(conn, table_name) for table_name in REQUIRED_TABLES)


def schema_needs_migration(conn) -> bool:
    if get_current_schema_version(conn) < SCHEMA_VERSION:
        return True
    if any(not column_exists(conn, "users", name) for name in REQUIRED_USER_COLUMNS):
        return True
    if any(not column_exists(conn, "records", name) for name in REQUIRED_RECORD_COLUMNS):
        return True
    return False


def validate_core_schema(conn):
    missing_user_columns = [name for name in REQUIRED_USER_COLUMNS if not column_exists(conn, "users", name)]
    missing_record_columns = [name for name in REQUIRED_RECORD_COLUMNS if not column_exists(conn, "records", name)]
    if missing_user_columns or missing_record_columns:
        raise RuntimeError(
            "Database schema is incomplete: "
            f"users missing {missing_user_columns or 'none'}, "
            f"records missing {missing_record_columns or 'none'}"
        )


def database_has_user_data(conn) -> bool:
    if not table_exists(conn, "users"):
        return False
    row = conn.execute("SELECT COUNT(*) FROM users").fetchone()
    return bool(row and row[0])


def backup_database(source_conn, label: str):
    BACKUP_DIR.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_path = BACKUP_DIR / f"bookkeeping_{label}_{timestamp}.db"
    target_conn = sqlite3.connect(str(backup_path))
    try:
        source_conn.backup(target_conn)
    finally:
        target_conn.close()

    wal_path = DB_PATH.with_suffix(".db-wal")
    shm_path = DB_PATH.with_suffix(".db-shm")
    if wal_path.exists():
        shutil.copy2(wal_path, BACKUP_DIR / f"bookkeeping_{label}_{timestamp}.db-wal")
    if shm_path.exists():
        shutil.copy2(shm_path, BACKUP_DIR / f"bookkeeping_{label}_{timestamp}.db-shm")
    return backup_path


def maybe_backup_before_migration(conn, current_version: int):
    global MIGRATION_BACKUP_DONE
    if MIGRATION_BACKUP_DONE:
        return None
    if current_version >= SCHEMA_VERSION:
        return None
    if not DB_PATH.exists():
        return None
    if current_version == 0 and not database_has_user_data(conn):
        return None
    backup_path = backup_database(conn, f"pre_v{SCHEMA_VERSION}")
    MIGRATION_BACKUP_DONE = True
    return backup_path


def prepare_database(conn):
    create_base_tables(conn)
    seed_users(conn)
    current_version = get_current_schema_version(conn)
    maybe_backup_before_migration(conn, current_version)
    migrate_schema(conn)
    validate_core_schema(conn)


def ensure_db_ready(conn):
    create_base_tables(conn)
    if has_required_tables(conn) and not schema_needs_migration(conn):
        validate_core_schema(conn)
        return


# ========== 初始化 / 迁移 ==========

def init_db():
    conn = get_conn()
    conn.close()


def create_base_tables(conn):
    conn.execute("""
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT UNIQUE NOT NULL,
            password_hash TEXT NOT NULL,
            display_name TEXT DEFAULT '',
            is_admin INTEGER DEFAULT 0,
            organization_id INTEGER REFERENCES organizations(id),
            role TEXT DEFAULT 'staff',
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS sessions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL REFERENCES users(id),
            token TEXT UNIQUE NOT NULL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            expires_at TIMESTAMP NOT NULL
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS organizations (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            owner_user_id INTEGER REFERENCES users(id),
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS records (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL DEFAULT 1 REFERENCES users(id),
            organization_id INTEGER REFERENCES organizations(id),
            created_by_user_id INTEGER REFERENCES users(id),
            date TEXT DEFAULT '',
            project_name TEXT DEFAULT '',
            applicant TEXT DEFAULT '',
            payer TEXT DEFAULT '',
            receiver TEXT DEFAULT '',
            purpose TEXT DEFAULT '',
            income REAL DEFAULT 0,
            expense REAL DEFAULT 0,
            refundable TEXT DEFAULT '否',
            expected_refund TEXT DEFAULT '',
            status TEXT DEFAULT '',
            remark TEXT DEFAULT '',
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS audit_events (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER REFERENCES users(id),
            username TEXT DEFAULT '',
            role TEXT DEFAULT '',
            organization_id INTEGER REFERENCES organizations(id),
            event_type TEXT NOT NULL,
            request_path TEXT DEFAULT '',
            detail_json TEXT DEFAULT '',
            prompt_tokens INTEGER DEFAULT 0,
            completion_tokens INTEGER DEFAULT 0,
            total_tokens INTEGER DEFAULT 0,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS settings (
            key TEXT PRIMARY KEY,
            value TEXT
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS schema_migrations (
            version INTEGER PRIMARY KEY,
            applied_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            note TEXT DEFAULT ''
        )
    """)


def get_current_schema_version(conn) -> int:
    row = conn.execute("SELECT COALESCE(MAX(version), 0) FROM schema_migrations").fetchone()
    return int(row[0] or 0)


def set_schema_version(conn, version: int, note: str):
    conn.execute(
        "INSERT OR REPLACE INTO schema_migrations (version, note) VALUES (?, ?)",
        (version, note),
    )


def migrate_schema(conn):
    version = get_current_schema_version(conn)
    if version < 1:
        migrate_to_v1(conn)
        version = 1
    if version < 2:
        migrate_to_v2(conn)
        version = 2
    if version < 3:
        migrate_to_v3(conn)
        version = 3
    if version < 4:
        migrate_to_v4(conn)


def migrate_to_v1(conn):
    ensure_column(conn, "records", "user_id", "user_id INTEGER NOT NULL DEFAULT 1 REFERENCES users(id)")
    set_schema_version(conn, 1, "base multi-user schema")


def migrate_to_v2(conn):
    ensure_column(conn, "users", "organization_id", "organization_id INTEGER REFERENCES organizations(id)")
    ensure_column(conn, "users", "role", "role TEXT DEFAULT 'staff'")
    ensure_column(conn, "records", "organization_id", "organization_id INTEGER REFERENCES organizations(id)")
    ensure_column(conn, "records", "created_by_user_id", "created_by_user_id INTEGER REFERENCES users(id)")
    set_schema_version(conn, 2, "organization and role foundation")


def migrate_to_v3(conn):
    seed_default_organization(conn)
    backfill_user_roles(conn)
    backfill_record_ownership(conn)
    set_schema_version(conn, 3, "organization ownership backfill")


def migrate_to_v4(conn):
    ensure_column(conn, "records", "owner_user_id", "owner_user_id INTEGER REFERENCES users(id)")
    backfill_record_owner_user(conn)
    set_schema_version(conn, 4, "record owner scope")


def seed_default_organization(conn):
    users = conn.execute(
        "SELECT id, username, display_name, is_admin, organization_id FROM users ORDER BY id"
    ).fetchall()
    if not users:
        return

    org_row = conn.execute("SELECT id FROM organizations ORDER BY id LIMIT 1").fetchone()
    default_org_id = org_row[0] if org_row else None
    if default_org_id is None:
        admin_row = next((row for row in users if row["username"] == "admin"), users[0])
        org_name = DEFAULT_ORGANIZATION_NAME
        cur = conn.execute(
            "INSERT INTO organizations (name, owner_user_id) VALUES (?, ?)",
            (org_name, admin_row["id"]),
        )
        default_org_id = cur.lastrowid

    for user in users:
        if not user["organization_id"]:
            conn.execute(
                "UPDATE users SET organization_id = ? WHERE id = ?",
                (default_org_id, user["id"]),
            )


def backfill_user_roles(conn):
    rows = conn.execute("SELECT id, username, is_admin, role FROM users").fetchall()
    for row in rows:
        role = row["role"]
        if role in {ROLE_ADMIN, ROLE_BOSS, ROLE_STAFF}:
            continue
        if row["is_admin"]:
            new_role = ROLE_ADMIN
        elif row["username"] == "admin":
            new_role = ROLE_BOSS
        else:
            new_role = ROLE_STAFF
        conn.execute("UPDATE users SET role = ? WHERE id = ?", (new_role, row["id"]))


def backfill_record_ownership(conn):
    rows = conn.execute(
        "SELECT r.id, r.user_id, r.organization_id, r.created_by_user_id, u.organization_id AS user_org_id "
        "FROM records r LEFT JOIN users u ON u.id = r.user_id"
    ).fetchall()
    for row in rows:
        organization_id = row["organization_id"] or row["user_org_id"]
        created_by_user_id = row["created_by_user_id"] or row["user_id"]
        conn.execute(
            "UPDATE records SET organization_id = ?, created_by_user_id = ? WHERE id = ?",
            (organization_id, created_by_user_id, row["id"]),
        )


def backfill_record_owner_user(conn):
    rows = conn.execute(
        "SELECT id, owner_user_id, created_by_user_id, user_id FROM records"
    ).fetchall()
    for row in rows:
        owner_user_id = row["owner_user_id"] or row["user_id"] or row["created_by_user_id"]
        conn.execute(
            "UPDATE records SET owner_user_id = ? WHERE id = ?",
            (owner_user_id, row["id"]),
        )


# ========== 密码 ==========

def hash_password(password: str) -> str:
    salt = "jizhang2026"
    return hashlib.sha256((salt + password).encode()).hexdigest()


def verify_password(password: str, password_hash: str) -> bool:
    return hash_password(password) == password_hash


# ========== 用户 ==========

def seed_users(conn=None):
    owns_conn = conn is None
    if owns_conn:
        conn = get_conn()
    existing = conn.execute("SELECT COUNT(*) FROM users").fetchone()[0]
    if existing == 0:
        conn.execute(
            "INSERT INTO users (username, password_hash, display_name, is_admin, role) VALUES (?, ?, ?, ?, ?)",
            ("admin", hash_password("admin888"), "开发人员", 1, ROLE_ADMIN),
        )
        conn.execute(
            "INSERT INTO users (username, password_hash, display_name, role) VALUES (?, ?, ?, ?)",
            ("boss_demo", hash_password("123456"), "演示老板", ROLE_BOSS),
        )
        conn.execute(
            "INSERT INTO users (username, password_hash, display_name, role) VALUES (?, ?, ?, ?)",
            ("staff_demo", hash_password("123456"), "演示员工", ROLE_STAFF),
        )
    if owns_conn:
        conn.commit()
        conn.close()


def authenticate(username: str, password: str) -> dict | None:
    conn = get_conn()
    row = conn.execute("SELECT * FROM users WHERE username = ?", (username,)).fetchone()
    conn.close()
    if not row or not verify_password(password, row["password_hash"]):
        return None
    return dict(row)


def get_user(user_id: int) -> dict | None:
    conn = get_conn()
    row = conn.execute(
        "SELECT id, username, display_name, is_admin, role, organization_id FROM users WHERE id = ?",
        (user_id,),
    ).fetchone()
    conn.close()
    return dict(row) if row else None


def get_user_with_password(user_id: int) -> dict | None:
    conn = get_conn()
    row = conn.execute(
        "SELECT id, username, password_hash, display_name, is_admin, role, organization_id FROM users WHERE id = ?",
        (user_id,),
    ).fetchone()
    conn.close()
    return dict(row) if row else None


def normalize_role(role: str | None, default: str = ROLE_STAFF) -> str:
    candidate = (role or default or ROLE_STAFF).strip().lower()
    if candidate not in VALID_ROLES:
        raise ValueError(f"invalid role: {role}")
    return candidate


def list_users(organization_id=None) -> list:
    conn = get_conn()
    if organization_id is None:
        rows = conn.execute(
            "SELECT id, username, display_name, is_admin, role, organization_id, created_at FROM users ORDER BY id"
        ).fetchall()
    else:
        rows = conn.execute(
            "SELECT id, username, display_name, is_admin, role, organization_id, created_at "
            "FROM users WHERE organization_id = ? ORDER BY id",
            (organization_id,),
        ).fetchall()
    conn.close()
    return [dict(row) for row in rows]


def create_user(username: str, password: str, display_name: str, organization_id=None, role: str = ROLE_STAFF) -> dict:
    role = normalize_role(role)
    def operation(conn):
        cur = conn.execute(
            "INSERT INTO users (username, password_hash, display_name, is_admin, organization_id, role) VALUES (?, ?, ?, ?, ?, ?)",
            (
                username,
                hash_password(password),
                display_name,
                int(role == ROLE_ADMIN),
                organization_id,
                role,
            ),
        )
        return cur.lastrowid

    user_id = run_write(operation)
    return get_user(user_id)


def update_user_password(user_id: int, new_password: str) -> bool:
    if not new_password or not str(new_password).strip():
        raise ValueError("password required")
    changed = run_write(
        lambda conn: conn.execute(
            "UPDATE users SET password_hash = ? WHERE id = ?",
            (hash_password(new_password), user_id),
        ).rowcount
    )
    return bool(changed)


def get_user_by_username(username: str) -> dict | None:
    conn = get_conn()
    row = conn.execute(
        "SELECT id, username, display_name, is_admin, role, organization_id FROM users WHERE username = ?",
        (username,),
    ).fetchone()
    conn.close()
    return dict(row) if row else None


def create_organization(name: str, owner_user_id=None) -> dict:
    conn = get_conn()
    cur = conn.execute(
        "INSERT INTO organizations (name, owner_user_id) VALUES (?, ?)",
        (name, owner_user_id),
    )
    organization_id = cur.lastrowid
    conn.close()
    return get_organization(organization_id)


def assign_user_organization(user_id: int, organization_id: int, role: str | None = None, is_admin=None) -> dict | None:
    if role is not None:
        role = normalize_role(role)
    def operation(conn):
        if role is None and is_admin is None:
            conn.execute(
                "UPDATE users SET organization_id = ? WHERE id = ?",
                (organization_id, user_id),
            )
        elif is_admin is None:
            conn.execute(
                "UPDATE users SET organization_id = ?, role = ? WHERE id = ?",
                (organization_id, role, user_id),
            )
        elif role is None:
            conn.execute(
                "UPDATE users SET organization_id = ?, is_admin = ? WHERE id = ?",
                (organization_id, int(bool(is_admin)), user_id),
            )
        else:
            conn.execute(
                "UPDATE users SET organization_id = ?, role = ?, is_admin = ? WHERE id = ?",
                (organization_id, role, int(bool(is_admin)), user_id),
            )

    run_write(operation)
    return get_user(user_id)


def get_organization_summary(organization_id: int) -> dict | None:
    organization = get_organization(organization_id)
    if not organization:
        return None
    members = list_users(organization_id)
    return {
        "organization": organization,
        "members": members,
    }


def update_organization_name(organization_id: int, name: str) -> dict | None:
    clean_name = (name or "").strip() or DEFAULT_ORGANIZATION_NAME
    run_write(
        lambda conn: conn.execute(
            "UPDATE organizations SET name = ? WHERE id = ?",
            (clean_name, organization_id),
        )
    )
    return get_organization(organization_id)


def normalize_organization_names():
    conn = get_conn()
    rows = conn.execute("SELECT id, name FROM organizations").fetchall()
    for row in rows:
        name = (row["name"] or "").strip()
        if not name or "?" in name:
            conn.execute(
                "UPDATE organizations SET name = ? WHERE id = ?",
                (DEFAULT_ORGANIZATION_NAME, row["id"]),
            )
    conn.close()


def user_has_related_records(user_id: int) -> bool:
    conn = get_conn()
    count = conn.execute(
        "SELECT COUNT(*) FROM records WHERE owner_user_id = ? OR created_by_user_id = ? OR user_id = ?",
        (user_id, user_id, user_id),
    ).fetchone()[0]
    conn.close()
    return count > 0


def delete_user(user_id: int) -> bool:
    def operation(conn):
        conn.execute("DELETE FROM sessions WHERE user_id = ?", (user_id,))
        cur = conn.execute("DELETE FROM users WHERE id = ?", (user_id,))
        return cur.rowcount > 0

    deleted = run_write(operation)
    return deleted


def can_manage_system(user: dict | None) -> bool:
    if not user:
        return False
    return bool(user.get("is_admin")) or user.get("role") == ROLE_ADMIN


def can_manage_organization_records(user: dict | None) -> bool:
    if not user:
        return False
    return can_manage_system(user) or user.get("role") == ROLE_BOSS


def build_record_scope(user: dict) -> tuple[str, list]:
    if can_manage_system(user):
        return "1 = 1", []
    organization_id = user.get("organization_id")
    if organization_id is None:
        return "1 = 0", []
    if can_manage_organization_records(user):
        return "organization_id = ?", [organization_id]
    return "organization_id = ? AND owner_user_id = ?", [organization_id, user["id"]]


def log_audit_event(
    event_type: str,
    *,
    user: dict | None = None,
    request_path: str = "",
    detail: dict | None = None,
    prompt_tokens: int = 0,
    completion_tokens: int = 0,
    total_tokens: int = 0,
):
    username = user.get("username", "") if user else ""
    role = user.get("role", "") if user else ""
    organization_id = user.get("organization_id") if user else None
    user_id = user.get("id") if user else None
    detail_json = json_dumps(detail or {})

    run_write(
        lambda conn: conn.execute(
            """
            INSERT INTO audit_events (
                user_id, username, role, organization_id, event_type, request_path,
                detail_json, prompt_tokens, completion_tokens, total_tokens
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                user_id,
                username,
                role,
                organization_id,
                event_type,
                request_path,
                detail_json,
                int(prompt_tokens or 0),
                int(completion_tokens or 0),
                int(total_tokens or 0),
            ),
        )
    )


def json_dumps(payload: dict) -> str:
    import json
    return json.dumps(payload, ensure_ascii=False)


def build_report_record_scope(user: dict, view: str | None = None, owner_user_id=None) -> tuple[str, list, dict]:
    normalized_view = (view or "").strip().lower()
    organization_id = user.get("organization_id")
    owner_user_id = int(owner_user_id) if owner_user_id not in (None, "") else None

    if can_manage_system(user):
        if owner_user_id is not None:
            return "owner_user_id = ?", [owner_user_id], {
                "view": "owner",
                "owner_user_id": owner_user_id,
                "base_balance_user_id": owner_user_id,
                "balance_label": "个人余额",
            }
        return "1 = 1", [], {
            "view": "all",
            "owner_user_id": None,
            "base_balance_user_id": user["id"],
            "balance_label": "当前余额",
        }

    if organization_id is None:
        return "1 = 0", [], {
            "view": "none",
            "owner_user_id": None,
            "base_balance_user_id": None,
            "balance_label": "当前余额",
        }

    if not can_manage_organization_records(user):
        return "organization_id = ? AND owner_user_id = ?", [organization_id, user["id"]], {
            "view": "mine",
            "owner_user_id": user["id"],
            "base_balance_user_id": user["id"],
            "balance_label": "当前余额",
        }

    if owner_user_id is not None:
        return "organization_id = ? AND owner_user_id = ?", [organization_id, owner_user_id], {
            "view": "owner",
            "owner_user_id": owner_user_id,
            "base_balance_user_id": owner_user_id,
            "balance_label": "个人余额",
        }

    if normalized_view == "team":
        return "organization_id = ? AND owner_user_id != ?", [organization_id, user["id"]], {
            "view": "team",
            "owner_user_id": None,
            "base_balance_user_id": None,
            "balance_label": "团队净额",
        }

    if normalized_view == "all":
        return "organization_id = ?", [organization_id], {
            "view": "all",
            "owner_user_id": None,
            "base_balance_user_id": None,
            "balance_label": "组织净额",
        }

    return "organization_id = ? AND owner_user_id = ?", [organization_id, user["id"]], {
        "view": "mine",
        "owner_user_id": user["id"],
        "base_balance_user_id": user["id"],
        "balance_label": "当前余额",
    }


def list_organizations() -> list:
    conn = get_conn()
    rows = conn.execute(
        """
        SELECT
            o.id,
            o.name,
            o.owner_user_id,
            o.created_at,
            COUNT(u.id) AS member_count
        FROM organizations o
        LEFT JOIN users u ON u.organization_id = o.id
        GROUP BY o.id, o.name, o.owner_user_id, o.created_at
        ORDER BY o.id
        """
    ).fetchall()
    conn.close()
    return [dict(row) for row in rows]


def set_organization_owner(organization_id: int, owner_user_id: int | None) -> dict | None:
    run_write(
        lambda conn: conn.execute(
            "UPDATE organizations SET owner_user_id = ? WHERE id = ?",
            (owner_user_id, organization_id),
        )
    )
    return get_organization(organization_id)


def get_organization(organization_id: int) -> dict | None:
    conn = get_conn()
    row = conn.execute("SELECT * FROM organizations WHERE id = ?", (organization_id,)).fetchone()
    conn.close()
    return dict(row) if row else None


# ========== 会话 ==========

def create_session(user_id: int, days: int = 7) -> str:
    token = secrets.token_hex(32)
    run_write(
        lambda conn: conn.execute(
            "INSERT INTO sessions (user_id, token, expires_at) VALUES (?, ?, ?)",
            (user_id, token, datetime.now() + timedelta(days=days)),
        )
    )
    return token


def get_user_by_session(token: str) -> dict | None:
    conn = get_conn()
    row = conn.execute(
        "SELECT u.id, u.username, u.display_name, u.is_admin, u.role, u.organization_id FROM users u "
        "JOIN sessions s ON u.id = s.user_id "
        "WHERE s.token = ? AND s.expires_at > ?",
        (token, datetime.now()),
    ).fetchone()
    conn.close()
    return dict(row) if row else None


def delete_session(token: str):
    run_write(lambda conn: conn.execute("DELETE FROM sessions WHERE token = ?", (token,)))


def cleanup_sessions():
    run_write(lambda conn: conn.execute("DELETE FROM sessions WHERE expires_at < ?", (datetime.now(),)))


# ========== 设置 (按用户隔离) ==========

def _setting_key(user_id: int, key: str) -> str:
    return f"{user_id}:{key}"


def set_setting(user_id: int, key: str, value: str):
    run_write(
        lambda conn: conn.execute(
            "INSERT OR REPLACE INTO settings (key, value) VALUES (?, ?)",
            (_setting_key(user_id, key), value),
        )
    )


def has_setting(user_id: int, key: str) -> bool:
    conn = get_conn()
    namespaced_key = _setting_key(user_id, key)
    row = conn.execute("SELECT 1 FROM settings WHERE key = ?", (namespaced_key,)).fetchone()
    if row:
        conn.close()
        return True
    legacy_row = conn.execute("SELECT value FROM settings WHERE key = ?", (key,)).fetchone()
    if legacy_row:
        conn.execute(
            "INSERT OR REPLACE INTO settings (key, value) VALUES (?, ?)",
            (namespaced_key, legacy_row["value"]),
        )
        conn.commit()
        conn.close()
        return True
    conn.close()
    return False


def get_setting(user_id: int, key: str, default=None):
    conn = get_conn()
    namespaced_key = _setting_key(user_id, key)
    row = conn.execute("SELECT value FROM settings WHERE key = ?", (namespaced_key,)).fetchone()
    if row:
        conn.close()
        return row["value"]
    legacy_row = conn.execute("SELECT value FROM settings WHERE key = ?", (key,)).fetchone()
    if legacy_row:
        conn.execute(
            "INSERT OR REPLACE INTO settings (key, value) VALUES (?, ?)",
            (namespaced_key, legacy_row["value"]),
        )
        conn.commit()
        conn.close()
        return legacy_row["value"]
    conn.close()
    return default


# ========== 记录 ==========

def save_records(records: list, user_id: int):
    user = get_user(user_id)
    organization_id = user.get("organization_id") if user else None
    def operation(conn):
        for rec in records:
            owner_user_id = rec.get("owner_user_id") or rec.get("归属人ID") or user_id
            conn.execute(
                """
                INSERT INTO records (
                    user_id, owner_user_id, organization_id, created_by_user_id, date, project_name, applicant, payer, receiver,
                    purpose, income, expense, refundable, expected_refund, status, remark
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    owner_user_id,
                    owner_user_id,
                    organization_id,
                    user_id,
                    rec.get("日期", ""),
                    rec.get("项目名称", ""),
                    rec.get("申请人", ""),
                    rec.get("转款人", ""),
                    rec.get("收款人", ""),
                    rec.get("用途", ""),
                    float(rec.get("收入", 0) or 0),
                    float(rec.get("支出", 0) or 0),
                    rec.get("可退回", "否"),
                    rec.get("预计退回", ""),
                    rec.get("状态", ""),
                    rec.get("备注", ""),
                ),
            )

    run_write(operation)


def update_record(record_id: int, data: dict):
    run_write(
        lambda conn: conn.execute(
            """
            UPDATE records SET
                date = ?, project_name = ?, applicant = ?, payer = ?, receiver = ?,
                purpose = ?, income = ?, expense = ?, refundable = ?,
                expected_refund = ?, status = ?, remark = ?
            WHERE id = ?
            """,
            (
                data.get("date", ""),
                data.get("project_name", ""),
                data.get("applicant", ""),
                data.get("payer", ""),
                data.get("receiver", ""),
                data.get("purpose", ""),
                float(data.get("income", 0) or 0),
                float(data.get("expense", 0) or 0),
                data.get("refundable", "否"),
                data.get("expected_refund", ""),
                data.get("status", ""),
                data.get("remark", ""),
                record_id,
            ),
        )
    )


def get_record_owner(record_id: int) -> int | None:
    conn = get_conn()
    row = conn.execute("SELECT owner_user_id FROM records WHERE id = ?", (record_id,)).fetchone()
    conn.close()
    return row["owner_user_id"] if row else None


def get_record(record_id: int) -> dict | None:
    conn = get_conn()
    row = conn.execute("SELECT * FROM records WHERE id = ?", (record_id,)).fetchone()
    conn.close()
    return dict(row) if row else None


def delete_records(record_ids: list[int], user: dict) -> int:
    if not record_ids:
        return 0
    placeholders = ",".join("?" for _ in record_ids)
    where_clause, params = build_record_scope(user)
    deleted = run_write(
        lambda conn: conn.execute(
            f"DELETE FROM records WHERE {where_clause} AND id IN ({placeholders})",
            params + record_ids,
        ).rowcount
    )
    return deleted


def get_balance(user: dict) -> float:
    init_bal = float(get_setting(user["id"], "initial_balance", "0") or 0)
    where_clause, params = build_record_scope(user)
    conn = get_conn()
    total_income = conn.execute(
        f"SELECT COALESCE(SUM(income), 0) FROM records WHERE {where_clause}",
        params,
    ).fetchone()[0]
    total_expense = conn.execute(
        f"SELECT COALESCE(SUM(expense), 0) FROM records WHERE {where_clause}",
        params,
    ).fetchone()[0]
    conn.close()
    return init_bal + total_income - total_expense


def get_records_with_balance(where_clause: str, params: list, base_balance: float = 0):
    conn = get_conn()
    rows = conn.execute(
        f"""
        SELECT * FROM records
        WHERE {where_clause}
        ORDER BY
            CASE WHEN date IS NULL OR date = '' THEN 1 ELSE 0 END,
            date,
            id
        """,
        params,
    ).fetchall()
    conn.close()
    result = []
    balance = base_balance
    for row in rows:
        balance = balance + (row["income"] or 0) - (row["expense"] or 0)
        record = dict(row)
        record["balance"] = balance
        result.append(record)
    return result


def get_balance_history(user: dict):
    init_bal = float(get_setting(user["id"], "initial_balance", "0") or 0)
    where_clause, params = build_record_scope(user)
    return get_records_with_balance(where_clause, params, init_bal)


def load_records(user: dict, limit=None):
    records = get_balance_history(user)
    if limit:
        records = records[-limit:]
    where_clause, params = build_record_scope(user)
    conn = get_conn()
    projects = [
        row[0]
        for row in conn.execute(
            f"SELECT DISTINCT project_name FROM records WHERE {where_clause} AND project_name != '' ORDER BY project_name",
            params,
        ).fetchall()
    ]
    conn.close()
    return {
        "records": records,
        "balance": get_balance(user),
        "projects": projects,
        "initial_balance_set": has_setting(user["id"], "initial_balance"),
    }


def has_records(user: dict) -> bool:
    where_clause, params = build_record_scope(user)
    conn = get_conn()
    count = conn.execute(f"SELECT COUNT(*) FROM records WHERE {where_clause}", params).fetchone()[0]
    conn.close()
    return count > 0


def get_all_records(user: dict):
    return {
        "records": get_balance_history(user),
        "balance": get_balance(user),
    }


def get_report_records(user: dict, view: str | None = None, owner_user_id=None) -> dict:
    where_clause, params, scope_meta = build_report_record_scope(user, view, owner_user_id)
    base_balance_user_id = scope_meta.get("base_balance_user_id")
    base_balance = 0
    if base_balance_user_id:
        base_balance = float(get_setting(base_balance_user_id, "initial_balance", "0") or 0)

    records = get_records_with_balance(where_clause, params, base_balance)
    total_income = sum((row.get("income") or 0) for row in records)
    total_expense = sum((row.get("expense") or 0) for row in records)
    projects = sorted({row.get("project_name") for row in records if row.get("project_name")})

    owner_options = []
    if user.get("organization_id") is not None and can_manage_organization_records(user):
        for member in list_users(user["organization_id"]):
            if member.get("is_admin") or member.get("role") == ROLE_ADMIN:
                continue
            owner_options.append({
                "id": member["id"],
                "label": member.get("display_name") or member.get("username"),
                "role": member.get("role"),
                "is_self": member["id"] == user["id"],
            })

    return {
        "records": records,
        "balance": base_balance + total_income - total_expense,
        "total_income": total_income,
        "total_expense": total_expense,
        "projects": projects,
        "view": scope_meta["view"],
        "owner_user_id": scope_meta.get("owner_user_id"),
        "balance_label": scope_meta.get("balance_label", "当前余额"),
        "owner_options": owner_options,
    }


def get_database_file_size() -> int:
    total = 0
    for path in (DB_PATH, DB_PATH.with_suffix(".db-wal"), DB_PATH.with_suffix(".db-shm")):
        if path.exists():
            total += path.stat().st_size
    return total


def get_admin_dashboard_snapshot() -> dict:
    conn = get_conn()

    counts = {
        "organizations": conn.execute("SELECT COUNT(*) FROM organizations").fetchone()[0],
        "users": conn.execute("SELECT COUNT(*) FROM users").fetchone()[0],
        "records": conn.execute("SELECT COUNT(*) FROM records").fetchone()[0],
        "sessions": conn.execute("SELECT COUNT(*) FROM sessions").fetchone()[0],
        "audit_events": conn.execute("SELECT COUNT(*) FROM audit_events").fetchone()[0],
    }

    role_counts = {
        "admins": conn.execute("SELECT COUNT(*) FROM users WHERE is_admin = 1 OR role = ?", (ROLE_ADMIN,)).fetchone()[0],
        "bosses": conn.execute("SELECT COUNT(*) FROM users WHERE role = ?", (ROLE_BOSS,)).fetchone()[0],
        "staff": conn.execute("SELECT COUNT(*) FROM users WHERE role = ?", (ROLE_STAFF,)).fetchone()[0],
    }

    token_usage = conn.execute(
        """
        SELECT
            COALESCE(SUM(prompt_tokens), 0) AS prompt_tokens,
            COALESCE(SUM(completion_tokens), 0) AS completion_tokens,
            COALESCE(SUM(total_tokens), 0) AS total_tokens
        FROM audit_events
        """
    ).fetchone()

    recent_activity = conn.execute(
        """
        SELECT id, username, role, event_type, request_path, total_tokens, created_at
        FROM audit_events
        ORDER BY id DESC
        LIMIT 20
        """
    ).fetchall()

    per_user_activity = conn.execute(
        """
        SELECT
            COALESCE(username, '') AS username,
            COALESCE(role, '') AS role,
            COUNT(*) AS event_count,
            COALESCE(SUM(total_tokens), 0) AS total_tokens,
            MAX(created_at) AS last_seen
        FROM audit_events
        GROUP BY COALESCE(username, ''), COALESCE(role, '')
        ORDER BY event_count DESC, last_seen DESC
        LIMIT 20
        """
    ).fetchall()

    record_distribution = conn.execute(
        """
        SELECT
            u.id,
            u.username,
            u.display_name,
            u.role,
            COUNT(r.id) AS record_count
        FROM users u
        LEFT JOIN records r ON r.owner_user_id = u.id
        GROUP BY u.id, u.username, u.display_name, u.role
        ORDER BY record_count DESC, u.id ASC
        LIMIT 20
        """
    ).fetchall()

    recent_records = conn.execute(
        """
        SELECT
            r.id,
            r.date,
            r.project_name,
            r.applicant,
            r.receiver,
            r.purpose,
            r.income,
            r.expense,
            r.owner_user_id,
            u.username,
            u.display_name,
            u.role,
            o.name AS organization_name
        FROM records r
        LEFT JOIN users u ON u.id = r.owner_user_id
        LEFT JOIN organizations o ON o.id = r.organization_id
        ORDER BY r.id DESC
        LIMIT 20
        """
    ).fetchall()

    conn.close()

    return {
        "counts": counts,
        "role_counts": role_counts,
        "token_usage": {
            "prompt_tokens": token_usage["prompt_tokens"],
            "completion_tokens": token_usage["completion_tokens"],
            "total_tokens": token_usage["total_tokens"],
        },
        "database_size_bytes": get_database_file_size(),
        "recent_activity": [dict(row) for row in recent_activity],
        "per_user_activity": [dict(row) for row in per_user_activity],
        "record_distribution": [dict(row) for row in record_distribution],
        "recent_records": [dict(row) for row in recent_records],
    }


def export_to_excel(user: dict):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    columns = ["日期", "项目名称", "申请人", "转款人", "收款人", "用途", "收入", "支出", "余额", "可退回", "预计退回", "状态", "备注"]
    money_fmt = '#,##0.00'

    wb = Workbook()
    ws = wb.active
    ws.title = "记账本"

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    for col_idx, name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.font = Font(bold=True, size=11, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    widths = [12, 14, 10, 10, 16, 20, 12, 12, 14, 8, 12, 8, 20]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    init_bal = float(get_setting(user["id"], "initial_balance", "0") or 0)
    ws.cell(row=2, column=1, value="初始余额")
    balance_cell = ws.cell(row=2, column=9, value=init_bal)
    balance_cell.number_format = money_fmt

    records = get_balance_history(user)
    for i, rec in enumerate(records):
        row = i + 3
        ws.cell(row=row, column=1, value=rec.get("date", ""))
        ws.cell(row=row, column=2, value=rec.get("project_name", ""))
        ws.cell(row=row, column=3, value=rec.get("applicant", ""))
        ws.cell(row=row, column=4, value=rec.get("payer", ""))
        ws.cell(row=row, column=5, value=rec.get("receiver", ""))
        ws.cell(row=row, column=6, value=rec.get("purpose", ""))
        c7 = ws.cell(row=row, column=7, value=rec.get("income") or 0)
        c7.number_format = money_fmt
        c8 = ws.cell(row=row, column=8, value=rec.get("expense") or 0)
        c8.number_format = money_fmt
        c9 = ws.cell(row=row, column=9, value=rec.get("balance") or 0)
        c9.number_format = money_fmt
        ws.cell(row=row, column=10, value=rec.get("refundable", "否"))
        ws.cell(row=row, column=11, value=rec.get("expected_refund", ""))
        ws.cell(row=row, column=12, value=rec.get("status", ""))
        ws.cell(row=row, column=13, value=rec.get("remark", ""))

    export_dir = DATA_DIR / "exports"
    export_dir.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    path = export_dir / f"bookkeeping_{ts}.xlsx"
    wb.save(path)
    return str(path)


init_db()
try:
    cleanup_sessions()
    normalize_organization_names()
except sqlite3.OperationalError:
    pass
